import openpyxl
import re
import os

def buscar_posicion_por_nombre(sheet, nombre_columna):
    for columna in sheet.iter_cols(min_row=1, max_row=1):
        for celda in columna:
            if celda.value.strip() == nombre_columna.strip():
                return celda.column  # Devuelve el número de columna (1-based index)
    return None  # Si no se encuentra, devuelve None

def procesar_archivo(file_path):
    wb = openpyxl.load_workbook(file_path)
    ws = wb.active

    # Buscar la posición de la columna TELEFONO_ALU
    columna_telefonos = buscar_posicion_por_nombre(ws, 'TELEFONO_ALU')
    if columna_telefonos is not None:
        # Insertar las nuevas columnas TELEFONO_1 y TELEFONO_2
        ws.insert_cols(columna_telefonos + 1, amount=2)
        ws.cell(row=1, column=columna_telefonos + 1, value='TELEFONO_1')
        ws.cell(row=1, column=columna_telefonos + 2, value='TELEFONO_2')

        # Procesar cada fila para dividir los números de teléfono
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=columna_telefonos, max_col=columna_telefonos):
            telefonos = row[0].value
            if telefonos:
                telefonos = str(telefonos)
                # Reemplazar múltiples signos de + por un solo +
                telefonos = re.sub(r'\++', '+', telefonos)
                # Si el número no tiene un signo de +, añadir uno al inicio
                if '+' not in telefonos:
                    telefonos = '+' + telefonos
                # Eliminar signos de + al inicio
                telefonos = re.sub(r'^\++', '', telefonos)
                numeros = telefonos.split('+')
                telefono_1 = numeros[0] if len(numeros) > 0 else ''
                telefono_2 = numeros[1] if len(numeros) > 1 else ''
                ws.cell(row=row[0].row, column=columna_telefonos + 1, value=telefono_1)
                ws.cell(row=row[0].row, column=columna_telefonos + 2, value=telefono_2)

    # Buscar la posición de la columna CORREO_ALU
    columna_correos = buscar_posicion_por_nombre(ws, 'CORREO_ALU')
    if columna_correos is not None:
        # Insertar las nuevas columnas CORREO_1 a CORREO_7
        ws.insert_cols(columna_correos + 1, amount=7)
        for i in range(1, 8):
            ws.cell(row=1, column=columna_correos + i, value=f'CORREO_{i}')

        # Procesar cada fila para dividir los correos electrónicos
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=columna_correos, max_col=columna_correos):
            correos = row[0].value
            if correos:
                # Reemplazar múltiples signos de + por un solo +
                correos = re.sub(r'\++', '+', correos)
                # # Eliminar signos de + al inicio
                # correos = re.sub(r'^\++', '', correos)
                emails = correos.split('+')
                for i in range(7):
                    ws.cell(row=row[0].row, column=columna_correos + 1 + i, value=emails[i] if i < len(emails) else '')

    # Añadir la nueva columna SECUE y calcular valores secuenciales
    emplid_dict = {}
    ws.insert_cols(3)  # Insertar una nueva columna en la posición C (columna 3)
    ws.cell(row=1, column=3, value='SECUE')  # Añadir cabecera SECUE

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
        emplid = row[4].value  # Cambia el índice según la columna EMPLID (índice 5, 0-based)
        if emplid not in emplid_dict:
            emplid_dict[emplid] = 0
        emplid_dict[emplid] += 1
        row[2].value = emplid_dict[emplid]  # Asignar el valor secuencial en la columna C (índice 3, 0-based)

    # Llenar la columna NUM_MASIVA con "EMPLID-DNI" si está vacía
    columna_num_masiva = buscar_posicion_por_nombre(ws, 'NUM_MASIVA')
    columna_emplid = buscar_posicion_por_nombre(ws, 'EMPLID')
    columna_dni = buscar_posicion_por_nombre(ws, 'DNI')
    if columna_num_masiva is not None and columna_emplid is not None and columna_dni is not None:
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            num_masiva = row[columna_num_masiva - 1].value  # -1 porque row es 0-based
            emplid = row[columna_emplid - 1].value  # -1 porque row es 0-based
            dni = row[columna_dni - 1].value  # -1 porque row es 0-based
            if not num_masiva:
                row[columna_num_masiva - 1].value = f'{emplid}-{dni}'

    # Guardar el archivo modificado con sufijo 'MOD' al inicio
    base_name, ext = os.path.splitext(file_path)
    directory = os.path.dirname(file_path)
    new_file_name = f"MOD_{os.path.basename(base_name)}{ext}"
    new_file_path = os.path.join(directory, new_file_name)
    wb.save(new_file_path)
    print(f'Archivo procesado y guardado como: {new_file_path}')

def procesar_archivos_en_directorio(directorio):
    for archivo in os.listdir(directorio):
        if archivo.endswith('.xlsx'):
            file_path = os.path.join(directorio, archivo)
            procesar_archivo(file_path)

# Ejecutar el script en el directorio actual
directorio_actual = os.getcwd()
procesar_archivos_en_directorio(directorio_actual)
